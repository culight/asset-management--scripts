import openpyxl as OPX
import sys
import getopt
import os.path
import shutil
from os.path import expanduser

input = ''
excel_file = ''

home_path = expanduser("~") + '/'

multi_dir = True

working_path = ''
strings_path = ''
excel_file_path = ''
replaced_path = ''
source_path = ''

source_files = []

current_file_path = ''
current_file_name = ''
current_file_rel_path = ''

def_exten = '.png'

def main(argv):
    global input
    global excel_file

    try:
        opts, args = getopt.getopt(argv, "hi:e:s", ["ifile=", "efile="])
    except getopt.GetoptError:
        print 'image_injection.py -i <working directory> -e <excel file name> ' \
              'e.g. python image_injection.py -i Users/user/project-path/extracted-assets/images/ -e image_excel_file'
        sys.exit(2)
    for opt, arg in opts:
        if opt == '-h':
            print 'image_injection.py -i <working directory> -e <excel file name> ' \
                  'e.g. python image_injection.py -i Users/user/project-path/extracted-assets/images/ -e image_excel_file'
            sys.exit()
        elif opt in ("-i", "--ifile"):
            input = arg
        elif opt in ("-e", "--ofile"):
            excel_file = arg


def initial_setup():
    global working_path
    global strings_path

    working_path = get_working_path()
    images_path = get_image_path()


def image_injection():
    global source_files

    print 'Checking image directories'
    for directory, dirnames, filenames in os.walk(strings_path):
        # If this path contains the necessary elements, set up the source and replaced folders and the excel object
        if len(filenames) > 0 and filenames[0].endswith('.xlsx'):
            directory_setup(directory, filenames)

        # If the source folder is not empty, start image injection
        if os.path.basename(directory) == 'source':
            source_files = os.listdir(directory)
            inject(source_files)

def inject(src):
    global current_file_name
    global current_file_path
    global current_file_rel_path
    global source_path
    global replaced_path

    # Open a read-only instance of the excel workbook
    wb_read = OPX.load_workbook(filename=excel_file_path, use_iterators=True)
    sheet_names_rd = wb_read.get_sheet_names()
    ws_read = wb_read.get_sheet_by_name(sheet_names_rd[0])

    # Open a write-only instance of the excel workbook
    # wb_write = OPX.load_workbook(filename=excel_file_path)
    # sheet_names_wr = wb_write.get_sheet_names()
    # ws_write = wb_write.get_sheet_by_name(sheet_names_wr[0])

    # Define appropriate range for iteration
    start_row = 2
    end_row = ws_read.max_row + 1
    iter_range = "{col}{start_row}:{col}{end_row}".format(col='E', start_row=start_row, end_row=end_row)

    # For each spreadsheet, iterate through its rows checking the current and source cells
    for sheet in sheet_names_rd:
        ws_read = wb_read.get_sheet_by_name(sheet)
        for row in ws_read.iter_rows(range_string=iter_range):
            if len(row) > 0:
                cell = row[0]
                if str(cell.value) != 'None':
                    index = cell.row
                    explicit_source_name = str(ws_read.cell('F' + str(index)).value)
                    if explicit_source_name != 'None':
                        source_file_name = explicit_source_name
                        get_current_path(str(ws_read.cell('E' + str(index)).value))
                        if source_file_name == 'remove':
                            remove_image()
                        elif source_file_name is not '':
                            swap_images(source_file_name,src)
                    else:
                        implicit_file_name = str(ws_read.cell('A' + str(index)).value)
                        source_file_name = implicit_file_name
                        get_current_path(str(ws_read.cell('E' + str(index)).value))
                        swap_images(source_file_name, src)
                else:
                    continue


def swap_images(src_name, src_list):
    if len(src_list) > 0:
        for image in src_list:
            if image == src_name:
                current_file = current_file_path
                source_file = source_path + src_name

                # Replace the current image with source image
                shutil.move(current_file, replaced_path)
                shutil.move(source_file, current_file_rel_path)

                src_list.remove(image)

                print current_file_name + ' has been replaced with ' + src_name
                return
            else:
                continue


def get_current_path(path):
    global current_file_path
    global current_file_name
    global current_file_rel_path

    current_file_path = path
    current_file_name = os.path.basename(current_file_path)
    current_file_rel_path = os.path.dirname(current_file_path) + '/'


def remove_image():
    global current_file_path
    global current_file_name

    current_file = current_file_path

    # Remove the current image
    shutil.move(current_file, replaced_path)

    print current_file_name + ' has been removed'


def directory_setup(dir, names):
    global replaced_path
    global source_path
    global excel_file_path

    excel_file_name = names[0]
    excel_file_path = dir + '/' + excel_file_name
    replaced_path = dir + '/replaced/'
    source_path = dir + '/source/'

    if not os.path.exists(replaced_path):
        sys.exit('replaced path does not exist')
    if not os.path.exists(source_path):
        sys.exit('source path does not exist')


def get_image_path():
    if os.path.exists(working_path + 'images'):
        path = working_path + 'images/'
    else:
        sys.exit('images directory does not exist')
    return path


def get_working_path():
    if home_path_present() is not True:
        abs_path = home_path + input
    else:
        abs_path = input
    return abs_path


def home_path_present():
    if home_path not in input:
        return False
    return True

if __name__ == "__main__":
    main(sys.argv[1:])

initial_setup()
if multi_dir is True:
    image_injection()